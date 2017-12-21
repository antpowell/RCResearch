import os
from openpyxl import Workbook
from openpyxl import load_workbook
import re

wb = Workbook()
ws = wb.active
ws.title = "Professors Directory List"
macroBookName = "RollCall-FirstStop.xlsm!"
getSMSMacroName = "begin"
attendanceCompareMacroName = "UpdateRoll.UpdateRoll"



def getProfessorList(folderName):
    numberOfProfessors = 0
    numberOfCourses = 0

    folders = os.listdir(folderName)
    i = 1
    for folder in folders:
        if re.search('\.', folder):
            print('true')
        else:
            numberOfProfessors += 1
            ws["A%d" % i] = folder
            files = os.listdir("D:\Google Drive\Mobile App Development\\2017 Fall Classes\Active\%s" % folder)
            if len(files) > 0:
                for file in files:
                    if re.search('(\.ini)', file):
                        print(file)
                    else:
                        numberOfCourses += 1
                        i += 1
                        ws["B%d" % i] = file
                        ws["C%d" % i] = "D:\Google Drive\Mobile App Development\\2017 Fall Classes\Active\{}\{}".format(folder, file)

            i += 1
    ws["A%d" % i] = "# of professors"
    ws["B%d" % i] = numberOfProfessors
    ws["C%d" % i] = "# of courses"
    ws["D%d" % i] = numberOfCourses
    # wb._archive.close()
    wb.save('ProfessorsList.xlsx')

getProfessorList("D:\Google Drive\Mobile App Development\\2017 Fall Classes\Active")