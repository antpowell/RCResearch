from pprint import pprint
import json
import openpyxl

xl_filepath = 'C:\\Users\powel\Desktop\Roll Call\Collection.xlsx'
wb = openpyxl.load_workbook(xl_filepath)
Full_Course_Student_List_sheet = wb.get_sheet_by_name('Full_Course_Student_List')

course_roster_json_file = open(
    'C:\\Users\powel\Desktop\Roll Call\python3\RollCall\course_roster.json', 'r', encoding='utf-8')
course_info_json_file = open(
    'C:\\Users\powel\Desktop\Roll Call\python3\RollCall\\full_course_information.json', 'r',
    encoding='utf-8')
course_roster = json.load(course_roster_json_file)
course_info = json.load(course_info_json_file)


start_row = 2
course_code_col = 1
instructor = 13
course_code = 14
course_name = 15
section = 16
day = 17
time = 18
credits = 19
building = 20
room = 21
college = 22
department = 23
term = 24
crn = 25

while start_row != len(Full_Course_Student_List_sheet['A']):
    if Full_Course_Student_List_sheet.cell(row=start_row, column=course_code_col).value in course_info['courses']:
        Full_Course_Student_List_sheet.cell(row=start_row, column=instructor, value=course_info['courses'][Full_Course_Student_List_sheet.cell(row=start_row, column=course_code_col).value]['instructor'])
        Full_Course_Student_List_sheet.cell(row=start_row, column=course_code, value=course_info['courses'][Full_Course_Student_List_sheet.cell(row=start_row, column=course_code_col).value]['course_code'])
        Full_Course_Student_List_sheet.cell(row=start_row, column=course_name, value=course_info['courses'][Full_Course_Student_List_sheet.cell(row=start_row, column=course_code_col).value]['course_name'])
        Full_Course_Student_List_sheet.cell(row=start_row, column=section, value=course_info['courses'][Full_Course_Student_List_sheet.cell(row=start_row, column=course_code_col).value]['section'])
        Full_Course_Student_List_sheet.cell(row=start_row, column=day, value=course_info['courses'][Full_Course_Student_List_sheet.cell(row=start_row, column=course_code_col).value]['day'])
        Full_Course_Student_List_sheet.cell(row=start_row, column=time, value=course_info['courses'][Full_Course_Student_List_sheet.cell(row=start_row, column=course_code_col).value]['time'])
        Full_Course_Student_List_sheet.cell(row=start_row, column=credits, value=course_info['courses'][Full_Course_Student_List_sheet.cell(row=start_row, column=course_code_col).value]['credits'])
        Full_Course_Student_List_sheet.cell(row=start_row, column=building, value=course_info['courses'][Full_Course_Student_List_sheet.cell(row=start_row, column=course_code_col).value]['building'])
        Full_Course_Student_List_sheet.cell(row=start_row, column=room, value=course_info['courses'][Full_Course_Student_List_sheet.cell(row=start_row, column=course_code_col).value]['room'])
        Full_Course_Student_List_sheet.cell(row=start_row, column=college, value=course_info['courses'][Full_Course_Student_List_sheet.cell(row=start_row, column=course_code_col).value]['college'])
        Full_Course_Student_List_sheet.cell(row=start_row, column=department, value=course_info['courses'][Full_Course_Student_List_sheet.cell(row=start_row, column=course_code_col).value]['department'])
        Full_Course_Student_List_sheet.cell(row=start_row, column=term, value=course_info['courses'][Full_Course_Student_List_sheet.cell(row=start_row, column=course_code_col).value]['term'])
        Full_Course_Student_List_sheet.cell(row=start_row, column=crn, value=course_info['courses'][Full_Course_Student_List_sheet.cell(row=start_row, column=course_code_col).value]['crn'])
    start_row += 1

# for row in Full_Course_Student_List_sheet.iter_rows(min_row=1, max_col=1):
#     for cell in row:
#         if cell.value in course_info['courses']:
#             print(row)
# Full_Course_Student_List_sheet['M{index}'] = course_info['courses'][cell.value]['instructor']
# Full_Course_Student_List_sheet['N{index}'] = course_info['courses'][cell.value]['course_code']
# Full_Course_Student_List_sheet['O{index}'] = course_info['courses'][cell.value]['course_name']
# Full_Course_Student_List_sheet['P{index}'] = course_info['courses'][cell.value]['section']
# Full_Course_Student_List_sheet['Q{index}'] = course_info['courses'][cell.value]['day']
# Full_Course_Student_List_sheet['R{index}'] = course_info['courses'][cell.value]['time']
# Full_Course_Student_List_sheet['S{index}'] = course_info['courses'][cell.value]['credits']
# Full_Course_Student_List_sheet['T{index}'] = course_info['courses'][cell.value]['building']
# Full_Course_Student_List_sheet['U{index}'] = course_info['courses'][cell.value]['room']
# Full_Course_Student_List_sheet['V{index}'] = course_info['courses'][cell.value]['college']
# Full_Course_Student_List_sheet['W{index}'] = course_info['courses'][cell.value]['department']
# Full_Course_Student_List_sheet['X{index}'] = course_info['courses'][cell.value]['term']
# Full_Course_Student_List_sheet['Y{index}'] = course_info['courses'][cell.value]['crn']
# print(course_info['courses'][cell.value])
wb.save('Collection.xlsx')

# pprint/.pprint(course_roster['course_rosters']['ENG 130 01'])
